"""
oscn_search.py
--------------
Searches OSCN (Oklahoma State Courts Network) by party/business name,
scraping results and appending new cases to oscn_cases.xlsx.

Launches a real visible Chrome window automatically (same approach as the
Vermont automation). If OSCN shows a Cloudflare challenge, the script
handles it automatically using human-like mouse movements.

Requirements
------------
    pip install selenium webdriver-manager openpyxl pyautogui

RUN
---
    python oscn_search.py

OUTPUT
------
oscn_cases.xlsx  (same folder as this script)
Columns: County | Case Number | Date Filed | Style
- Deduplicates by Case Number against any existing rows in the file.
- Alternating row shading, frozen header, auto-filter, running Total row.
"""

import time
import random
import math
from datetime import datetime, timedelta
from pathlib import Path

import pyautogui
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
#  CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

TARGET_URL  = "https://www.oscn.net/dockets/Search.aspx#all"
XLSX_PATH   = Path(__file__).parent / "oscn_cases.xlsx"
DAYS_BACK   = 90

LAST_NAMES = [
    "JG Wentworth",
    "J.G. Wentworth",
    "J G Wentworth",
    "J. G. Wentworth",
    "JG Wentworth Organizations",
    "J.G. Wentworth Organizations",
    "JG Wentworth Organizations, LLC",
    "J.G. Wentworth Organizations, LLC",
    "J G Wentworth Organizations, LLC",
    "J. G. Wentworth Organizations, LLC",
    "JG Wentworth Organizations LLC",
    "J.G. Wentworth Organizations LLC",
    "J G Wentworth Organizations LLC",
    "J. G. Wentworth Organizations LLC",
    "DRB Capital LLC",
    "DRB Capital, LLC",
    "Stone Street Capital LLC",
    "Stone Street Capital, LLC",
    "AA Ron I LLC",
    "AA Ron I, LLC",
    "Abactor LLC",
    "Abactor, LLC",
    "Abidole LLC",
    "Abidole, LLC",
    "Adenna Med LLC",
    "Adenna Med, LLC",
    "Adventura LLC",
    "Adventura, LLC",
    "AGPI LLC",
    "AGPI, LLC",
    "Aikman Structured Finance LLC",
    "Aikman Structured Finance, LLC",
    "Annuity Transfers Ltd",
    "Apis Management LLC",
    "Apis Management, LLC",
    "Atlas Legal Funding III LP",
    "AXE Finance LLC",
    "AXE Finance, LLC",
    "B.A.W.21",
    "B.R. Wright LLC",
    "B.R. Wright, LLC",
    "BHG Structured Settlements Inc",
    "BHG Structured Settlements, Inc",
    "Bifco, LLC",
    "Blue Grape LLC",
    "Blue Grape, LLC",
    "Catalina Structured Funding Inc",
    "Catalina Structured Funding, Inc",
    "Concordis Group Limited",
    "Conrad Factoring, LLC",
    "Conrad Factoring LLC",
    "Cornerstone Funding LLC",
    "Cornerstone Funding, LLC",
    "Fast Annuity Settlement Transfers LLC",
    "Fast Annuity Settlement Transfers, LLC",
    "FL Assignments Corp",
    "G.D.T.R.F.B. LLC",
    "G.D.T.R.F.B., LLC",
    "G7 Crescenta LLC",
    "G7 Crescenta, LLC",
    "Genex Capital Corp",
    "GJ 123 LLC",
    "GJ 123, LLC",
    "Greenwood Funding LLC",
    "Greenwood Funding, LLC",
    "Grier I LLC",
    "Grier I, LLC",
    "Hakstol Group LLC",
    "Hakstol Group, LLC",
    "Hiddenview Ent, LLC",
    "JLC Capital Funding, LLC",
    "KN Direct Capital LLC",
    "KN Direct Capital, LLC",
    "Lane Nimitz LLC",
    "Lane Nimitz, LLC",
    "Lasko LLC",
    "Lasko, LLC",
    "Leaf 002 LLC",
    "Leaf 002, LLC",
    "Legere LLC",
    "Legere, LLC",
    "Lottery Funding, LLC",
    "M McDougall LLC",
    "M McDougall, LLC",
    "Majestic Funding LLC",
    "Majestic Funding, LLC",
    "Mic-Bry8",
    "Olive Branch Funding LLC",
    "Olive Branch Funding, LLC",
    "Palermo Group LLC",
    "Palermo Group, LLC",
    "Palm Green Closing, LLC",
    "Palm Harbor LLC",
    "Palm Harbor, LLC",
    "Passira Mal LLC",
    "Passira Mal, LLC",
    "Patriot Settlement Resources LLC",
    "Patriot Settlement Resources, LLC",
    "QLS Funding LLC",
    "QLS Funding, LLC",
    "Reliance Funding LLC",
    "Reliance Funding, LLC",
    "Rocorp Corporation",
    "RSL Funding LLC",
    "RSL Funding, LLC",
    "Savannah Settlements LLC",
    "Savannah Settlements, LLC",
    "Sempra Finance LLC",
    "Sempra Finance, LLC",
    "Seneca Originations",
    "SeneOne LLC",
    "SeneOne, LLC",
    "Settlement Capital Corp",
    "Settlement Status LLC",
    "Settlement Status, LLC",
    "Somerton LLC",
    "Somerton, LLC",
    "Stratcap Investments Inc.",
    "Stratcap Investments, Inc.",
    "Stratton Asset Funding LLC",
    "Stratton Asset Funding, LLC",
    "Structured Asset Funding",
    "TKD LLC",
    "TKD, LLC",
    "TRM V LLC",
    "TRM V, LLC",
    "Tybenz LLC",
    "Tybenz, LLC",
    "Uber Funding LLC",
    "Uber Funding, LLC",
    "Vintage Equity Group LLC",
    "Vintage Equity Group, LLC",
    "Wepaymore Funding",
    "Wepaymore Funding, LLC",
    "Zakho Way LLC",
    "Zakho Way, LLC",
    "GREAT PLAINS MANAGEMENT CORPORATION",
    "T ENE LLC",
    "T ENE, LLC",
    "RD FITZ LLC",
    "RD FITZ, LLC",
    "GA OFF LLC",
    "GA OFF, LLC",
    "Assured Management Corporation",
    "BENTZEN FINANCIAL LLC",
    "BENTZEN FINANCIAL, LLC",
]

# ─────────────────────────────────────────────────────────────────────────────
#  HUMAN-LIKE MOUSE MOVEMENT
# ─────────────────────────────────────────────────────────────────────────────

def bezier_point(p0, p1, p2, p3, t):
    """Calculate a point on a cubic Bezier curve at parameter t (0-1)."""
    x = ((1-t)**3 * p0[0] + 3*(1-t)**2*t * p1[0] +
         3*(1-t)*t**2 * p2[0] + t**3 * p3[0])
    y = ((1-t)**3 * p0[1] + 3*(1-t)**2*t * p1[1] +
         3*(1-t)*t**2 * p2[1] + t**3 * p3[1])
    return (int(x), int(y))


def human_move_and_click(target_x, target_y):
    """
    Move the mouse to (target_x, target_y) along a natural Bezier curve,
    then click. Mimics how a human hand moves — curved path, variable speed,
    slight overshoot and wobble.
    """
    start_x, start_y = pyautogui.position()

    # Random control points create a natural curved arc
    cp1 = (
        start_x + random.randint(-80, 80),
        start_y + random.randint(-60, 60)
    )
    cp2 = (
        target_x + random.randint(-60, 60),
        target_y + random.randint(-60, 60)
    )

    # Number of steps — more steps = smoother movement
    steps = random.randint(40, 70)

    for i in range(steps + 1):
        t = i / steps
        # Ease in/out: slow at start and end, fast in middle
        t_eased = t * t * (3 - 2 * t)
        px, py = bezier_point((start_x, start_y), cp1, cp2, (target_x, target_y), t_eased)

        # Add tiny random wobble to simulate hand tremor
        px += random.randint(-1, 1)
        py += random.randint(-1, 1)

        pyautogui.moveTo(px, py, duration=0)

        # Variable speed: faster in middle, slower near start/end
        if i < steps * 0.2 or i > steps * 0.8:
            time.sleep(random.uniform(0.008, 0.015))   # slow
        else:
            time.sleep(random.uniform(0.002, 0.006))   # fast

    # Small pause before clicking (human hesitation)
    time.sleep(random.uniform(0.05, 0.15))

    # Click with slight randomness within the target element
    pyautogui.click(
        target_x + random.randint(-3, 3),
        target_y + random.randint(-3, 3)
    )


def get_element_center(driver, element):
    """Get the screen coordinates of the center of a Selenium element."""
    # Get element position relative to page
    rect = driver.execute_script("""
        var r = arguments[0].getBoundingClientRect();
        return {left: r.left, top: r.top, width: r.width, height: r.height};
    """, element)

    # Get browser window position on screen
    win_x = driver.execute_script("return window.screenX || window.screenLeft;")
    win_y = driver.execute_script("return window.screenY || window.screenTop;")

    # Account for browser chrome (toolbar height ~90px on most systems)
    toolbar_height = 90

    screen_x = win_x + rect['left'] + rect['width'] / 2
    screen_y = win_y + toolbar_height + rect['top'] + rect['height'] / 2

    return int(screen_x), int(screen_y)


# ─────────────────────────────────────────────────────────────────────────────
#  BROWSER
# ─────────────────────────────────────────────────────────────────────────────

def build_driver() -> webdriver.Chrome:
    """Launch a real, visible Chrome window."""
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    service = Service(ChromeDriverManager().install())
    driver  = webdriver.Chrome(service=service, options=options)
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return driver


def js(driver, script, *args):
    return driver.execute_script(script, *args)


# ─────────────────────────────────────────────────────────────────────────────
#  CLOUDFLARE HANDLER
# ─────────────────────────────────────────────────────────────────────────────

def handle_cloudflare(driver):
    """
    Detect and handle the Cloudflare Turnstile challenge.

    Three scenarios:
      1. No challenge — do nothing.
      2. Auto-verified (shows "Success!") — click Submit immediately.
      3. Manual checkbox on page — move mouse naturally to checkbox,
         click it, wait for Success, then click Submit.
    """
    MAX_WAIT = 60
    start    = time.time()

    while True:
        title        = driver.title.lower()
        is_turnstile = "turnstile" in title
        is_blocked   = any(x in title for x in
                           ["just a moment", "attention required", "cloudflare"])

        if not (is_turnstile or is_blocked):
            break   # normal page — nothing to do

        if time.time() - start > MAX_WAIT:
            print("   ⚠️  Cloudflare wait timed out — continuing anyway")
            break

        print(f"   🔒 Cloudflare page detected: '{driver.title}'")

        # Check if already showing Success (auto-verified)
        success_shown = driver.execute_script("""
            var body = document.body.innerText || '';
            return body.toLowerCase().includes('success');
        """)
        token_ready = driver.execute_script("""
            var inp = document.querySelector('[name="cf-turnstile-response"]');
            return inp && inp.value && inp.value.length > 10;
        """)

        if success_shown or token_ready:
            print("   ✅ Cloudflare verified (Success shown) — clicking Submit...")
            try:
                submit = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR,
                        'input[value="Submit"], button[type="submit"], '
                        'input[type="submit"], button'
                    ))
                )
                submit.click()
                print("   ✅ Submit clicked — waiting for results page...")
                WebDriverWait(driver, 15).until(
                    lambda d: "turnstile" not in d.title.lower()
                              and "just a moment" not in d.title.lower()
                )
                print(f"   ✅ Passed Cloudflare — now on: {driver.title}")
            except Exception as e:
                print(f"   ⚠️  Could not click Submit: {e}")
            break

        # ── Click checkbox using visual position (shadow DOM blocks Selenium) ──
        # The Cloudflare widget is inside a closed shadow root so Selenium
        # cannot find any elements inside it. We use fixed visual coordinates
        # based on the known page layout (maximized Chrome, 1920x1200 screen).
        checkbox_clicked = False
        try:
            # Get actual toolbar height (outerHeight - innerHeight)
            win_x     = driver.execute_script("return window.screenX || window.screenLeft;")
            win_y     = driver.execute_script("return window.screenY || window.screenTop;")
            outer_h   = driver.execute_script("return window.outerHeight;")
            inner_h   = driver.execute_script("return window.innerHeight;")
            outer_w   = driver.execute_script("return window.outerWidth;")
            toolbar_h = outer_h - inner_h

            # Cloudflare widget is centered horizontally on the page
            # Checkbox is ~120px left of the page center
            page_center_x = win_x + outer_w // 2
            checkbox_x    = page_center_x - 120 + random.randint(-3, 3)

            # Widget appears ~230px from top of content area
            checkbox_y    = win_y + toolbar_h + 234 + random.randint(-3, 3)

            print(f"   🖱️  Clicking checkbox at ({checkbox_x}, {checkbox_y}) "
                  f"[toolbar={toolbar_h}px, win=({win_x},{win_y})]")

            # Wander mouse around first (looks more human)
            cur_x, cur_y = pyautogui.position()
            for _ in range(random.randint(2, 4)):
                pyautogui.moveTo(
                    cur_x + random.randint(-200, 200),
                    cur_y + random.randint(-150, 150),
                    duration=random.uniform(0.3, 0.7)
                )
                time.sleep(random.uniform(0.1, 0.4))

            # Move to checkbox with natural Bezier curve and click
            human_move_and_click(checkbox_x, checkbox_y)
            checkbox_clicked = True
            print("   🖱️  Checkbox area clicked — waiting for verification...")
            time.sleep(4)

        except Exception as e:
            print(f"   ⚠️  Error clicking checkbox: {e}")

        # ── Fallback: check inside iframes ────────────────────────────────────
        if not checkbox_clicked:
            try:
                iframes = driver.find_elements(By.TAG_NAME, "iframe")
                for iframe in iframes:
                    src = iframe.get_attribute("src") or ""
                    if "challenges.cloudflare.com" in src or "turnstile" in src:
                        driver.switch_to.frame(iframe)
                        try:
                            checkbox = WebDriverWait(driver, 3).until(
                                EC.presence_of_element_located(
                                    (By.CSS_SELECTOR,
                                     "input[type='checkbox'], .ctp-checkbox-label, #cf-stage")
                                )
                            )
                            driver.switch_to.default_content()
                            iframe_x, iframe_y = get_element_center(driver, iframe)
                            print("   🖱️  Cloudflare checkbox found in iframe...")

                            cur_x, cur_y = pyautogui.position()
                            for _ in range(2):
                                pyautogui.moveTo(
                                    cur_x + random.randint(-150, 150),
                                    cur_y + random.randint(-100, 100),
                                    duration=random.uniform(0.3, 0.6)
                                )
                                time.sleep(random.uniform(0.1, 0.3))

                            human_move_and_click(iframe_x, iframe_y)
                            checkbox_clicked = True
                            print("   🖱️  Checkbox clicked — waiting for verification...")
                            time.sleep(3)
                            break
                        except Exception:
                            driver.switch_to.default_content()
                            continue
            except Exception as e:
                driver.switch_to.default_content()
                print(f"   ⚠️  Error looking in iframes: {e}")

        if not checkbox_clicked:
            print("   ⏳ Waiting for Cloudflare verification...")

        time.sleep(3)


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

SHEET_NAME = "OSCN Results"
COL_WIDTHS = {"County": 22, "Case Number": 20, "Date Filed": 15, "Style": 58}
HEADERS    = list(COL_WIDTHS.keys())

_HDR_FONT  = Font(name="Arial", bold=True, size=11, color="FFFFFFFF")
_HDR_FILL  = PatternFill("solid", fgColor="2F5496")
_HDR_ALIGN = Alignment(vertical="center", horizontal="center")
_HDR_BORD  = Border(bottom=Side(style="medium", color="1F3864"))

_ROW_FONT  = Font(name="Arial", size=10)
_ROW_ALIGN = Alignment(vertical="center")
_ROW_BORD  = Border(bottom=Side(style="thin", color="BFBFBF"))
_TOT_FONT  = Font(name="Arial", bold=True, size=10)
_TOT_ALIGN = Alignment(horizontal="center")

_FILL_EVEN = PatternFill("solid", fgColor="DCE6F1")
_FILL_ODD  = PatternFill("solid", fgColor="FFFFFF")


def load_existing_case_numbers() -> set:
    if not XLSX_PATH.exists():
        return set()
    try:
        wb = openpyxl.load_workbook(XLSX_PATH)
        if SHEET_NAME not in wb.sheetnames:
            return set()
        ws = wb[SHEET_NAME]
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[1]
            if val and not str(val).startswith("Total:"):
                existing.add(str(val).strip())
        print(f"  📂  {len(existing)} existing case(s) in {XLSX_PATH.name}")
        return existing
    except Exception as e:
        print(f"  [warn] Could not read existing file: {e}")
        return set()


def append_to_excel(new_cases: list):
    existing_row_count = 0

    if XLSX_PATH.exists():
        wb = openpyxl.load_workbook(XLSX_PATH)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
            last_row = ws.max_row
            if last_row >= 2:
                cell_val = str(ws.cell(row=last_row, column=3).value or "")
                if cell_val.startswith("Total:"):
                    ws.delete_rows(last_row)
            existing_row_count = ws.max_row - 1
        else:
            ws = _create_sheet(wb)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = _create_sheet(wb)

    for i, c in enumerate(new_cases):
        row_idx = ws.max_row + 1
        ws.cell(row=row_idx, column=1, value=c["county"])
        ws.cell(row=row_idx, column=2, value=c["caseNumber"])
        ws.cell(row=row_idx, column=3, value=c["dateFiled"])
        ws.cell(row=row_idx, column=4, value=c["style"])

        fill = _FILL_EVEN if (existing_row_count + i) % 2 == 0 else _FILL_ODD
        for col in range(1, 5):
            cell = ws.cell(row=row_idx, column=col)
            cell.font      = _ROW_FONT
            cell.fill      = fill
            cell.alignment = _ROW_ALIGN
            cell.border    = _ROW_BORD

    total_data_rows = existing_row_count + len(new_cases)
    total_row_idx   = ws.max_row + 1
    ws.cell(row=total_row_idx, column=3, value=f"Total: {total_data_rows} cases")
    ws.cell(row=total_row_idx, column=3).font      = _TOT_FONT
    ws.cell(row=total_row_idx, column=3).alignment = _TOT_ALIGN

    wb.save(XLSX_PATH)


def _create_sheet(wb: openpyxl.Workbook):
    ws = wb.create_sheet(SHEET_NAME)
    for col_idx, (header, width) in enumerate(COL_WIDTHS.items(), start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = _HDR_ALIGN
        cell.border    = _HDR_BORD
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    return ws


# ─────────────────────────────────────────────────────────────────────────────
#  SEARCH
# ─────────────────────────────────────────────────────────────────────────────

def _try_click(driver, selectors: list, label: str) -> bool:
    for sel in selectors:
        try:
            el = driver.find_element(By.CSS_SELECTOR, sel)
            if el.is_displayed():
                el.click()
                print(f"   ✔ {label} ({sel})")
                return True
        except (NoSuchElementException, Exception):
            continue
    return False


def _try_fill(driver, selectors: list, value: str, label: str) -> bool:
    for sel in selectors:
        try:
            el = driver.find_element(By.CSS_SELECTOR, sel)
            if el.is_displayed():
                el.click()
                time.sleep(0.2)
                el.clear()
                el.send_keys(value)
                print(f"   ✔ {label}: \"{value}\" ({sel})")
                return True
        except (NoSuchElementException, Exception):
            continue
    return False


def search_name(driver, last_name: str) -> list:
    print(f"\n{'─'*55}")
    print(f"🔎 Searching for: \"{last_name}\"")
    print(f"{'─'*55}")

    try:
        driver.get(TARGET_URL)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
    except TimeoutException:
        print("   ⚠️  Load timed out — page may still be rendering.")
    time.sleep(1)

    # Diagnostic — helps identify Cloudflare page title/URL when it appears
    print(f"   📄 Title: '{driver.title}' | URL: {driver.current_url[:70]}")

    handle_cloudflare(driver)

    # Check "All Databases"
    try:
        all_db_selectors = [
            'input[name="db"][value="all"]',
            'input#db_all',
            'input[value="all"]',
            'label[for="db_all"]',
            'a[href*="#all"]',
        ]
        checked = False
        for sel in all_db_selectors:
            try:
                el = driver.find_element(By.CSS_SELECTOR, sel)
                if el.is_displayed():
                    tag = el.tag_name.lower()
                    if tag == "input" and el.get_attribute("type") == "checkbox":
                        if not el.is_selected():
                            el.click()
                        print(f"   ✔ \"All Databases\" checkbox checked ({sel})")
                    else:
                        el.click()
                        print(f"   ✔ Clicked \"All Databases\" element ({sel})")
                    checked = True
                    break
            except (NoSuchElementException, Exception):
                continue

        if not checked:
            js(driver, """
                var f = document.querySelector('select[name="db"], input[name="db"]');
                if (f) { f.value = 'all'; f.dispatchEvent(new Event('change', {bubbles:true})); }
            """)
            print("   ⚠️  Checkbox not found — injected db=all via JS fallback")
    except Exception as e:
        print(f"   ⚠️  Error setting All Databases: {e}")

    time.sleep(0.3)

    filled = _try_fill(driver, ['#ln', 'input[name="lname"]'], last_name, "Last Name")
    if not filled:
        print("   ⚠️  Could not find Last Name field.")

    time.sleep(0.5)

    try:
        js(driver, "document.body.click();")
    except Exception:
        pass
    time.sleep(0.3)

    try:
        today     = datetime.today()
        prior     = today - timedelta(days=DAYS_BACK)
        fmt       = lambda d: d.strftime("%m/%d/%Y")
        date_from = fmt(prior)
        date_to   = fmt(today)
        print(f"   Date range: {date_from} → {date_to}")
        _try_fill(driver, ["#fdl"], date_from, "Start date")
        time.sleep(0.2)
        _try_fill(driver, ["#fdh"], date_to,   "End date")
    except Exception as e:
        print(f"   ⚠️  Error filling date range: {e}")

    time.sleep(0.5)

    submitted = _try_click(driver, [
        'input[type="submit"]',
        'button[type="submit"]',
        'input[value="Search"]',
        '#searchButton',
        'button',
    ], "Search submitted")
    if not submitted:
        print("   ⚠️  Could not find submit button.")

    time.sleep(2)

    # Cloudflare Turnstile sometimes appears after search submission
    # Title = 'OSCN Turnstile' — if Success shown, just click Submit
    if "turnstile" in driver.title.lower():
        print("   🔒 Cloudflare Turnstile detected after search — handling...")
        handle_cloudflare(driver)

    time.sleep(2)

    cases = js(driver, """
        var results = [];
        var currentCounty = "Unknown";
        var nodes = document.querySelectorAll(".caseCourtHeader, .resultTableRow");
        nodes.forEach(function(node) {
            if (node.classList.contains("caseCourtHeader")) {
                var raw   = node.innerText.trim();
                var match = raw.match(/(^.*?County)/i);
                currentCounty = match ? match[1] : raw;
                return;
            }
            function get(cls) {
                var el = node.querySelector("." + cls);
                return el ? el.innerText.trim() : "";
            }
            var caseNumber = get("result_casenumber");
            var dateFiled  = get("result_datefiled");
            var style      = get("result_shortstyle") || get("shortStyle");
            if (caseNumber) {
                results.push({
                    county:     currentCounty,
                    caseNumber: caseNumber,
                    dateFiled:  dateFiled,
                    style:      style
                });
            }
        });
        return results;
    """)

    cases = cases or []
    print(f"   📋 Found {len(cases)} result(s) for \"{last_name}\"")
    return cases


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("🚀 Launching Chrome...")
    driver = build_driver()

    print("🌐 Navigating to OSCN search page...")
    driver.get(TARGET_URL)
    time.sleep(3)

    # Diagnostic — shows exactly what title/URL Cloudflare uses when it appears
    print(f"   📄 Title: '{driver.title}' | URL: {driver.current_url[:70]}")

    handle_cloudflare(driver)

    print(f"✅ OSCN page accessible! Title: {driver.title}")

    existing_case_numbers = load_existing_case_numbers()
    print(f"📂 Existing cases in Excel: {len(existing_case_numbers)}")

    total_added = 0

    for last_name in LAST_NAMES:
        cases = search_name(driver, last_name)

        new_cases = [c for c in cases
                     if c["caseNumber"].strip() not in existing_case_numbers]
        dup_count = len(cases) - len(new_cases)

        if dup_count > 0:
            print(f"   ⏭️  Skipped {dup_count} duplicate(s)")

        if not new_cases:
            print(f"   ℹ️  No new cases to add for \"{last_name}\"")
            continue

        for c in new_cases:
            existing_case_numbers.add(c["caseNumber"].strip())

        append_to_excel(new_cases)
        total_added += len(new_cases)
        print(f"   💾 Appended {len(new_cases)} new case(s) to {XLSX_PATH.name}")

        time.sleep(1)

    print(f"\n✅ All done! Total new cases added this run: {total_added}")
    print(f"💾 Excel file: {XLSX_PATH}")
    print("🏁 Done.")


if __name__ == "__main__":
    main()